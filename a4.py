# Databricks notebook source
import os
import requests
import pandas as pd
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker
import sqlite3
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, PieChart, Reference
from textblob import TextBlob


# CONFIG
API_KEY = os.environ.get("YOUTUBE_API_KEY", "YOUR_API_KEY_HERE")
CHANNEL_ID = "UCYO_jab_esuFRV4b17AJtAw"  # 3Blue1Brown
SEARCH_URL = "https://www.googleapis.com/youtube/v3/search"
VIDEOS_URL = "https://www.googleapis.com/youtube/v3/videos"
N = 50
OUTPUT_EXCEL = "3b1b_analytics_report.xlsx"
OUTPUT_DB = "3b1b_analytics.db"


# Styling constants for Excel
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
ALT_FILL    = PatternFill("solid", fgColor="D6E4F0")
CENTER      = Alignment(horizontal="center", vertical="center")
THIN        = Side(style="thin", color="BFBFBF")
BORDER      = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


#1 fetch data from YouTube API

def fetch_videos(api_key, channel_id, n=50, order="date"):
    search_params = {
        "key": api_key,
        "channelId": channel_id,
        "part": "snippet",
        "order": order,
        "maxResults": n,
        "type": "video"
    }
    if api_key in (None, "YOUR_API_KEY_HERE"):
        raise ValueError("You must set a valid YOUTUBE_API_KEY in the environment or update API_KEY in the script.")

    search_response = requests.get(SEARCH_URL, params=search_params)
    try:
        search_response.raise_for_status()
    except requests.RequestException as exc:
        print(f"Step 1: Search request failed: {exc}")
        return pd.DataFrame(columns=["videoId", "title", "publishedAt", "viewCount", "likeCount", "commentCount"])

    search_data = search_response.json()
    video_ids = [item["id"]["videoId"] for item in search_data.get("items", [])]

    print(f"Step 1: Fetched {len(video_ids)} video IDs (ordered by {order})")
    if not video_ids:
        print("Step 1: No videos found. Check your channel ID, API key, or quota.")
        return pd.DataFrame(columns=["videoId", "title", "publishedAt", "viewCount", "likeCount", "commentCount"])

    video_params = {
        "key": api_key,
        "id": ",".join(video_ids),
        "part": "snippet,statistics"
    }
    video_response = requests.get(VIDEOS_URL, params=video_params)
    try:
        video_response.raise_for_status()
    except requests.RequestException as exc:
        print(f"Step 1: Video details request failed: {exc}")
        return pd.DataFrame(columns=["videoId", "title", "publishedAt", "viewCount", "likeCount", "commentCount"])

    video_data = video_response.json()

    video_list = []
    for item in video_data.get("items", []):
        snippet = item.get("snippet", {})
        stats = item.get("statistics", {})
        video_list.append({
            "videoId":      item.get("id"),
            "title":        snippet.get("title"),
            "publishedAt":  snippet.get("publishedAt"),
            "viewCount":    stats.get("viewCount", 0),
            "likeCount":    stats.get("likeCount", 0),
            "commentCount": stats.get("commentCount", 0)
        })

    df = pd.DataFrame(video_list)
    if df.empty:
        print("Step 1: No video details were returned for the fetched IDs.")
        return pd.DataFrame(columns=["videoId", "title", "publishedAt", "viewCount", "likeCount", "commentCount"])

    df["publishedAt"] = pd.to_datetime(df["publishedAt"])
    for col in ["viewCount", "likeCount", "commentCount"]:
        df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)

    print("Step 1: Complete\n")
    return df


#2 analyze data
def analyse(df):
    df = df.copy()

    # Engagement ratios
    df["likes_to_views_ratio"]    = (df["likeCount"]    / df["viewCount"]).fillna(0)
    df["comments_to_views_ratio"] = (df["commentCount"] / df["viewCount"]).fillna(0)

    # Sentiment on video titles
    def get_sentiment(text):
        polarity = TextBlob(text).sentiment.polarity
        if polarity > 0:   return "Positive"
        elif polarity < 0: return "Negative"
        else:              return "Neutral"

    df["sentiment"] = df["title"].apply(get_sentiment)

    # Print summary stats
    most_viewed   = df.loc[df["viewCount"].idxmax()]
    most_liked    = df.loc[df["likeCount"].idxmax()]
    most_commented = df.loc[df["commentCount"].idxmax()]

    print("Step 2: Analysis complete")
    print(f"  Most Viewed:    {most_viewed['title']} ({int(most_viewed['viewCount']):,} views)")
    print(f"  Most Liked:     {most_liked['title']} ({int(most_liked['likeCount']):,} likes)")
    print(f"  Most Commented: {most_commented['title']} ({int(most_commented['commentCount']):,} comments)")
    print(f"  Sentiment:      {df['sentiment'].value_counts().to_dict()}")
    print()

    return df


#3 visualize data, save pngs
def visualise(df):
    df = df.sort_values("publishedAt")

    #view count over time
    plt.figure(figsize=(12, 5))
    plt.plot(df["publishedAt"], df["viewCount"], marker="o", color="#1F4E79", linewidth=2)
    plt.title("View Count Over Time — 3Blue1Brown", fontsize=14)
    plt.xlabel("Publish Date")
    plt.ylabel("Views")
    plt.xticks(rotation=45)
    plt.gca().yaxis.set_major_formatter(mticker.StrMethodFormatter('{x:,.0f}'))
    plt.grid(alpha=0.4)
    plt.tight_layout()
    plt.savefig("chart_views_over_time.png", dpi=150)
    plt.close()

    # top 15 by like/view ratio
    df_sorted = df.nlargest(15, "likes_to_views_ratio")
    short_titles = [t[:35] + "..." if len(t) > 35 else t for t in df_sorted["title"]]
    fig, ax = plt.subplots(figsize=(12, 6))
    ax.barh(short_titles, df_sorted["likes_to_views_ratio"], color="#2E75B6")
    ax.set_xlabel("Likes / Views Ratio")
    ax.set_title("Top 15 Videos by Like-to-View Ratio — 3Blue1Brown", fontsize=13)
    ax.invert_yaxis()
    plt.tight_layout()
    plt.savefig("chart_like_ratio.png", dpi=150)
    plt.close()

    # top 10 most viewed
    top10 = df.nlargest(10, "viewCount")
    short_titles_top10 = [t[:35] + "..." if len(t) > 35 else t for t in top10["title"]]
    fig, ax = plt.subplots(figsize=(12, 6))
    ax.barh(short_titles_top10, top10["viewCount"], color="#1F4E79")
    ax.set_xlabel("View Count")
    ax.set_title("Top 10 Most Viewed Videos — 3Blue1Brown", fontsize=13)
    ax.xaxis.set_major_formatter(mticker.StrMethodFormatter('{x:,.0f}'))
    ax.invert_yaxis()
    plt.tight_layout()
    plt.savefig("chart_top10_views.png", dpi=150)
    plt.close()

    # sentiment pie
    sentiment_counts = df["sentiment"].value_counts()
    plt.figure(figsize=(6, 6))
    plt.pie(
        sentiment_counts,
        labels=sentiment_counts.index,
        autopct="%1.1f%%",
        colors=["#2E75B6", "#70AD47", "#ED7D31"]
    )
    plt.title("Sentiment Distribution of Video Titles — 3Blue1Brown")
    plt.tight_layout()
    plt.savefig("chart_sentiment.png", dpi=150)
    plt.close()

    print("Step 3: Charts saved as PNG files\n")

#4 export sqlite
def export_db(df, db_path=OUTPUT_DB):
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()

    cursor.executescript("""
        DROP TABLE IF EXISTS videos;
        CREATE TABLE videos (
            video_id                TEXT PRIMARY KEY,
            title                   TEXT,
            published_at            TEXT,
            view_count              INTEGER,
            like_count              INTEGER,
            comment_count           INTEGER,
            likes_to_views_ratio    REAL,
            comments_to_views_ratio REAL,
            sentiment               TEXT
        );
    """)

    df_db = df.copy()
    df_db["publishedAt"] = df_db["publishedAt"].astype(str)
    df_db = df_db.rename(columns={
        "videoId":               "video_id",
        "title":                 "title",
        "publishedAt":           "published_at",
        "viewCount":             "view_count",
        "likeCount":             "like_count",
        "commentCount":          "comment_count",
        "likes_to_views_ratio":  "likes_to_views_ratio",
        "comments_to_views_ratio": "comments_to_views_ratio",
        "sentiment":             "sentiment"
    })
    df_db.to_sql("videos", conn, if_exists="append", index=False)
    conn.commit()

    queries = [
        ("Top 10 Most Viewed Videos", "SELECT title, view_count, like_count, comment_count FROM videos ORDER BY view_count DESC LIMIT 10"), ("Top 10 by Like-to-View Ratio","SELECT title, view_count, ROUND(likes_to_views_ratio,4) as like_ratio FROM videos ORDER BY likes_to_views_ratio DESC LIMIT 10"),("Average Engagement by Sentiment","SELECT sentiment, COUNT(*) as count, ROUND(AVG(view_count),0) as avg_views, ROUND(AVG(likes_to_views_ratio),4) as avg_like_ratio FROM videos GROUP BY sentiment ORDER BY avg_views DESC"),("Videos with Above-Average Views","SELECT title, view_count FROM videos WHERE view_count > (SELECT AVG(view_count) FROM videos) ORDER BY view_count DESC"),
    ]

    for title_text, query in queries:
        print(f"\n--- {title_text} ---")
        print(pd.read_sql_query(query, conn).to_string(index=False))

    conn.close()
    print(f"\nStep 4: Database saved to {db_path}\n")


#5 export excel

def _style_header_row(ws, row_num, num_cols):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.font  = HEADER_FONT
        cell.fill  = HEADER_FILL
        cell.alignment = CENTER
        cell.border = BORDER

def _style_data_rows(ws, start_row, end_row, num_cols):
    for row in range(start_row, end_row + 1):
        fill = ALT_FILL if row % 2 == 0 else PatternFill()
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = fill
            cell.border = BORDER
            cell.alignment = Alignment(vertical="center")

def _write_df_to_sheet(ws, df_in, start_row=1):
    for r_idx, row in enumerate(dataframe_to_rows(df_in, index=False, header=True), start_row):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)
    _style_header_row(ws, start_row, len(df_in.columns))
    _style_data_rows(ws, start_row + 1, start_row + len(df_in), len(df_in.columns))
    return start_row + len(df_in) + 1

def _set_col_widths(ws, widths):
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width

def export_excel(df, output_path=OUTPUT_EXCEL):
    wb = openpyxl.Workbook()

    #Overview
    ws1 = wb.active
    ws1.title = "Overview"
    ws1.sheet_view.showGridLines = False
    ws1["B2"].value = "YouTube Analytics Report — 3Blue1Brown"
    ws1["B2"].font = Font(name="Calibri", bold=True, size=16, color="1F4E79")
    ws1.merge_cells("B2:G2")
    stats = [
        ("Channel",                 "3Blue1Brown (Mathematics)"),
        ("Videos Analysed",         len(df)),
        ("Date Range",              f"{df['publishedAt'].min().strftime('%Y-%m-%d')} → {df['publishedAt'].max().strftime('%Y-%m-%d')}"),
        ("Most Viewed Video",       df.loc[df['viewCount'].idxmax(), 'title']),
        ("Highest Like Ratio Video",df.loc[df['likes_to_views_ratio'].idxmax(), 'title']),
        ("Avg Views per Video",     f"{int(df['viewCount'].mean()):,}"),
        ("Avg Likes per Video",     f"{int(df['likeCount'].mean()):,}"),
    ]
    for i, (label, value) in enumerate(stats, 4):
        ws1.cell(row=i, column=2, value=label).font = Font(bold=True, color="1F4E79")
        ws1.cell(row=i, column=3, value=str(value))
    _set_col_widths(ws1, {"B": 26, "C": 60})

    #full data
    ws2 = wb.create_sheet("Video Data")
    ws2.sheet_view.showGridLines = False
    df_export = df[["videoId","title","publishedAt","viewCount","likeCount",
                    "commentCount","likes_to_views_ratio","comments_to_views_ratio","sentiment"]].copy()
    df_export["publishedAt"] = df_export["publishedAt"].dt.strftime("%Y-%m-%d")
    df_export["likes_to_views_ratio"]    = df_export["likes_to_views_ratio"].round(4)
    df_export["comments_to_views_ratio"] = df_export["comments_to_views_ratio"].round(4)
    df_export.columns = ["Video ID","Title","Publish Date","Views","Likes", "Comments","Like Ratio","Comment Ratio","Sentiment"]
    _write_df_to_sheet(ws2, df_export)
    _set_col_widths(ws2, {"A":14,"B":50,"C":14,"D":12,"E":10,"F":10,"G":12,"H":14,"I":12})

    # engagement analysis
    ws3 = wb.create_sheet("Engagement Analysis")
    ws3.sheet_view.showGridLines = False
    top_ratio = df.nlargest(15, "likes_to_views_ratio")[["title","viewCount","likeCount","likes_to_views_ratio"]].copy()
    top_ratio["likes_to_views_ratio"] = top_ratio["likes_to_views_ratio"].round(4)
    top_ratio.columns = ["Title","Views","Likes","Like Ratio"]
    ws3["A1"].value = "Top 15 Videos by Like-to-View Ratio"
    ws3["A1"].font = Font(bold=True, size=12, color="1F4E79")
    _write_df_to_sheet(ws3, top_ratio, start_row=2)
    _set_col_widths(ws3, {"A":55,"B":14,"C":12,"D":12})
    chart3 = BarChart()
    chart3.type = "bar"
    chart3.title = "Like-to-View Ratio (Top 15)"
    chart3.style = 10
    chart3.width = 22
    chart3.height = 14
    data3 = Reference(ws3, min_col=4, min_row=2, max_row=17)
    cats3 = Reference(ws3, min_col=1, min_row=3, max_row=17)
    chart3.add_data(data3, titles_from_data=True)
    chart3.set_categories(cats3)
    ws3.add_chart(chart3, "F2")
 
    # sentiment analysis
    ws4 = wb.create_sheet("Sentiment Analysis")
    ws4.sheet_view.showGridLines = False
    sent_table = df["sentiment"].value_counts().reset_index()
    sent_table.columns = ["Sentiment","Count"]
    sent_table["Percentage"] = (sent_table["Count"] / len(df) * 100).round(1).astype(str) + "%"
    _write_df_to_sheet(ws4, sent_table)
    _set_col_widths(ws4, {"A":14,"B":10,"C":12})
    pie = PieChart()
    pie.title = "Title Sentiment Distribution"
    pie.style = 10
    pie.width = 14
    pie.height = 10
    labels   = Reference(ws4, min_col=1, min_row=2, max_row=1+len(sent_table))
    data_pie = Reference(ws4, min_col=2, min_row=1, max_row=1+len(sent_table))
    pie.add_data(data_pie, titles_from_data=True)
    pie.set_categories(labels)
    ws4.add_chart(pie, "E2")
 
    # SQL results
    ws5 = wb.create_sheet("SQL Query Results")
    ws5.sheet_view.showGridLines = False
    conn = sqlite3.connect(OUTPUT_DB)
    queries = [
        ("Top 10 Most Viewed Videos",
         "SELECT title, view_count, like_count, comment_count FROM videos ORDER BY view_count DESC LIMIT 10"),
        ("Top 10 by Like-to-View Ratio",
         "SELECT title, view_count, ROUND(likes_to_views_ratio,4) as like_ratio FROM videos ORDER BY likes_to_views_ratio DESC LIMIT 10"),
        ("Average Engagement by Sentiment",
         "SELECT sentiment, COUNT(*) as count, ROUND(AVG(view_count),0) as avg_views, ROUND(AVG(likes_to_views_ratio),4) as avg_like_ratio FROM videos GROUP BY sentiment ORDER BY avg_views DESC"),
        ("Videos with Above-Average Views",
         "SELECT title, view_count FROM videos WHERE view_count > (SELECT AVG(view_count) FROM videos) ORDER BY view_count DESC"),
    ]
    current_row = 1
    for title_text, query in queries:
        ws5.cell(row=current_row, column=1, value=title_text).font = Font(bold=True, size=11, color="1F4E79")
        current_row += 1
        result_df = pd.read_sql_query(query, conn)
        current_row = _write_df_to_sheet(ws5, result_df, start_row=current_row)
        current_row += 1
    conn.close()
    _set_col_widths(ws5, {"A":55,"B":14,"C":14,"D":14})
 
    wb.save(output_path)
    print(f"Step 5: Excel report saved to {output_path}")
    print("Sheets: Overview, Video Data, Engagement Analysis, Sentiment Analysis, SQL Query Results\n")
    

# main
def main():
    print("=== 3Blue1Brown YouTube Analytics Pipeline ===\n")
    df = fetch_videos(API_KEY, CHANNEL_ID, n=N)
    if df.empty:
        print("Pipeline stopped: no video data available to process.")
        return

    df = analyse(df)
    visualise(df)
    export_db(df)
    export_excel(df)
    print("=== Pipeline complete ===")
 
if __name__ == "__main__":
    main()